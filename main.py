from __future__ import annotations

import math
import re
import shutil
import unicodedata
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import UnidentifiedImageError


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "DANE_WEJSCIOWE"
FALLBACK_INPUT_DIR = BASE_DIR / "DOCS"
OUTPUT_DIR = BASE_DIR / "GOTOWE_INSTRUKCJE"
TEMPLATE_PATH = BASE_DIR / "wzor.xlsx"
ASSETS_DIR = BASE_DIR / "assets"
TEMP_IMAGES_DIR = BASE_DIR / "temp_images"


# Jeśli przesuniesz pola w Excelu, zmień adresy komórek tutaj.
CELL_MAP = {
    "producer": "B8",
    "product_name": "D8",
    "revision_date": "F8",
    "hazards": "A11",
    "hand_protection": "B16",
    "respiratory_protection": "F16",
    "skin_protection": "B17",
    "eye_protection": "F17",
    "first_aid_inhalation": "A41",
    "first_aid_skin": "A42",
    "first_aid_eyes": "A43",
    "first_aid_ingestion": "A44",
    "first_aid_general": "A40",
    "fire_overview": "A48",
    "fire_suitable": "A49",
    "fire_unsuitable": "A50",
    "before_work_1": "A19",
    "before_work_2": "A20",
    "before_work_3": "A21",
    "before_work_4": "A22",
    "before_work_5": "A23",
    "before_work_6": "A24",
    "before_work_7": "A25",
    "before_work_8": "A26",
    "during_work_1": "C29",
    "during_work_2": "C30",
    "during_work_3": "C31",
    "during_work_4": "C32",
    "during_work_5": "C33",
    "during_work_6": "C34",
    "after_work_1": "A36",
    "after_work_2": "A37",
    "environmental_release": "A51",
    "handling_1": "A54",
    "storage_1": "A55",
}


# Pozycje ikon w arkuszu głównym. Jeśli w szablonie coś przesuniesz, zmień anchor.
STATIC_IMAGE_ANCHORS = {
    "gloves": "A16",
    "respirator": "E16",
    "clothing": "A17",
    "goggles": "E17",
    "no_smoking": "A29",
    "no_food": "A31",
}

HAZARD_IMAGE_ANCHORS = ["F11", "G11", "H11", "I11"]

APPEND_CELLS = {"A40", "A41", "A42", "A43", "A44", "A48", "A49", "A50", "A51", "A54", "A55"}


# Teksty stałe wpisywane zawsze (wg uwag: „dodać na stałe do wzoru instrukcji").
# Te komórki NIE są wypełniane danymi z karty — zawierają standardowe formułki.
FIRST_AID_GENERAL_TEXT = (
    "Okazać lekarzowi kartę charakterystyki lub etykietę produktu. Należy "
    "przestrzegać uwag dotyczących bezpieczeństwa i użytkowania zamieszczonych "
    "na etykiecie. Przy wystąpieniu symptomów lub przy wypadkach zasięgnąć rady "
    "lekarza. Usunąć poszkodowanego ze strefy zagrożenia. Jeśli poszkodowany "
    "jest nieprzytomny ułożyć w pozycji bocznej bezpiecznej, kontrolować oddech."
)
HANDLING_TEXT = (
    "Stosować środki ochrony indywidualnej. Zapewnić odpowiednią wentylację "
    "otoczenia. Nie jeść, nie pić, nie palić w czasie pracy. Umyć ręce po użyciu. "
    "Zanieczyszczoną odzież umyć przed ponownym użyciem."
)
STORAGE_TEXT = (
    "Przechowywać w oryginalnych, właściwie oznakowanych, szczelnie zamkniętych "
    "opakowaniach w chłodnym, suchym, dobrze wentylowanym miejscu. Trzymać z dala "
    "od produktów spożywczych."
)

# Wg uwag: czcionka 11 w sekcjach „Czynności przed rozpoczęciem pracy" (A19:A26)
# i „Czynności w czasie pracy" (C29:C34) — to statyczny tekst szablonu.
SECTION_FONT11_CELLS = [f"A{r}" for r in range(19, 27)] + [f"C{r}" for r in range(29, 35)]


GHS_LABELS = {
    "GHS01": "Wybuch",
    "GHS02": "Płomień",
    "GHS03": "Utleniacz",
    "GHS04": "Gaz pod ciśnieniem",
    "GHS05": "Żrący",
    "GHS06": "Toksyczny",
    "GHS07": "Drażniący / uczulający",
    "GHS08": "Poważne zagrożenie zdrowia",
    "GHS09": "Środowisko",
}


H_TO_GHS = {
    "H200": {"GHS01"},
    "H201": {"GHS01"},
    "H202": {"GHS01"},
    "H203": {"GHS01"},
    "H204": {"GHS01"},
    "H205": {"GHS01"},
    "H220": {"GHS02"},
    "H221": {"GHS02"},
    "H222": {"GHS02"},
    "H223": {"GHS02"},
    "H224": {"GHS02"},
    "H225": {"GHS02"},
    "H226": {"GHS02"},
    "H228": {"GHS02"},
    "H229": {"GHS04"},
    "H270": {"GHS03"},
    "H271": {"GHS03", "GHS05"},
    "H272": {"GHS03"},
    "H280": {"GHS04"},
    "H281": {"GHS04"},
    "H290": {"GHS05"},
    "H300": {"GHS06"},
    "H301": {"GHS06"},
    "H302": {"GHS07"},
    "H304": {"GHS08"},
    "H310": {"GHS06"},
    "H311": {"GHS06", "GHS07"},
    "H312": {"GHS07"},
    "H314": {"GHS05"},
    "H315": {"GHS07"},
    "H317": {"GHS07"},
    "H318": {"GHS05"},
    "H319": {"GHS07"},
    "H330": {"GHS06"},
    "H331": {"GHS06", "GHS07"},
    "H332": {"GHS07"},
    "H334": {"GHS08"},
    "H335": {"GHS07"},
    "H336": {"GHS07"},
    "H340": {"GHS08"},
    "H341": {"GHS08"},
    "H350": {"GHS08"},
    "H351": {"GHS08"},
    "H360": {"GHS08"},
    "H361": {"GHS08"},
    "H362": {"GHS08"},
    "H370": {"GHS08"},
    "H371": {"GHS08"},
    "H372": {"GHS08"},
    "H373": {"GHS08"},
    "H400": {"GHS09"},
    "H410": {"GHS09"},
    "H411": {"GHS09"},
    "H412": {"GHS09"},
    "H413": {"GHS09"},
}


@dataclass
class ExtractedData:
    product_name: str = ""
    producer: str = ""
    revision_date: str = ""
    hazard_text: str = ""
    hazard_codes: list[str] | None = None
    hazard_pictograms: list[str] | None = None
    hand_protection: str = ""
    respiratory_protection: str = ""
    skin_protection: str = ""
    eye_protection: str = ""
    first_aid_inhalation: str = ""
    first_aid_skin: str = ""
    first_aid_eyes: str = ""
    first_aid_ingestion: str = ""
    first_aid_general: str = ""
    fire_overview: str = ""
    fire_suitable: str = ""
    fire_unsuitable: str = ""
    environmental_release: str = ""
    before_work: list[str] | None = None
    during_work: list[str] | None = None
    after_work: list[str] | None = None
    storage: list[str] | None = None
    handling: list[str] | None = None
    # NDS (Najwyższe dopuszczalne stężenie) exposure limit description from sekcja 8.1
    nds: str = ""

    def __post_init__(self) -> None:
        self.hazard_codes = self.hazard_codes or []
        self.hazard_pictograms = self.hazard_pictograms or []
        self.before_work = self.before_work or []
        self.during_work = self.during_work or []
        self.after_work = self.after_work or []
        self.storage = self.storage or []
        self.handling = self.handling or []


# Znaki sterujące niedozwolone w XML/openpyxl (rzucają IllegalCharacterError
# przy zapisie komórki). Czasem trafiają do tekstu wyciągniętego z PDF/AI.
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean_xml(value):
    if isinstance(value, str):
        return _ILLEGAL_XML_RE.sub("", value)
    return value


def normalize_text(text: str) -> str:
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def compact_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def strip_accents(text: str) -> str:
    text = text.replace("ł", "l").replace("Ł", "L")
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def section_slice(text: str, start_patterns: Iterable[str], end_patterns: Iterable[str]) -> str:
    start = None
    for pattern in start_patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            start = match.start()
            break
    if start is None:
        return ""

    tail = text[start:]
    end_index = None
    for pattern in end_patterns:
        match = re.search(pattern, tail, flags=re.IGNORECASE)
        if match and match.start() > 0:
            end_index = match.start()
            break
    return tail[:end_index].strip() if end_index else tail.strip()


def capture_first(text: str, patterns: Iterable[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return normalize_text(match.group(1))
    return ""


def sanitize_filename(name: str) -> str:
    safe = strip_accents(name)
    safe = re.sub(r"[^\w\s.-]", "", safe)
    safe = re.sub(r"\s+", "_", safe.strip())
    return safe[:120] or "produkt"


def read_pdf_text(pdf_path: Path) -> str:
    pages: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                pages.append(page_text)
    return normalize_text("\n".join(pages))


def extract_product_name(text: str) -> str:
    value = capture_first(
        text,
        [
            r"1\.1[.\s]+Identyfikator produktu\s*(.+?)\s*1\.2",
            r"SEKCJA 1:.*?1\.1[.\s]+Identyfikator produktu\s*(.+?)\s*1\.2",
        ],
    )
    value = re.sub(r"^[·\s]*Nazwa handlowa:\s*", "", value, flags=re.IGNORECASE)
    value = re.split(r"\n[·\s]*UFI:", value, maxsplit=1, flags=re.IGNORECASE)[0]
    return normalize_text(value)


def extract_revision_date(text: str) -> str:
    return capture_first(
        text,
        [
            r"Aktualizacja(?: / numer wersji)?\s*:\s*([0-9]{2}[./-][0-9]{2}[./-][0-9]{4})",
            r"Obowiązuje od:\s*([0-9]{2}[./-][0-9]{2}[./-][0-9]{4})",
            r"Zmieniona wersja z dnia / numer wersji:\s*([0-9]{2}[./-][0-9]{2}[./-][0-9]{4})",
        ],
    )


def extract_producer(text: str) -> str:
    block = capture_first(
        text,
        [
            r"1\.3[.\s]+Dane dotyczące dostawcy karty charakterystyki\s*(.+?)\s*1\.4",
        ],
    )
    if not block:
        return ""

    lines = [line.strip() for line in block.splitlines() if line.strip()]
    filtered: list[str] = []
    for line in lines:
        line = re.sub(r"^[·\s]*", "", line)
        if re.search(r"(tel\.|fax|e-?mail|homepage|www\.|https?://|numer alarmowy)", line, re.IGNORECASE):
            continue
        if re.search(r"(komórka udzielająca informacji|komorka udzielajaca informacji|producent/dostawca:?)", line, re.IGNORECASE):
            continue
        filtered.append(line)
    return normalize_text("\n".join(filtered[:6]))


def extract_hazard_section(text: str) -> tuple[str, list[str], list[str]]:
    section2 = section_slice(text, [r"SEKCJA 2: Identyfikacja zagrożeń"], [r"SEKCJA 3"])
    if not section2:
        return "", [], []

    pictogram_codes = sorted(set(re.findall(r"\bGHS0[1-9]\b", section2)))
    # restrict hazard extraction to labeling subsection (2.2) to avoid classification duplicates
    label_section = section_slice(
        section2,
        [r"2\\.2[.\\s]+"],
        [r"2\\.3[.\\s]+", r"SEKCJA 3"]
    )

    hazard_block = capture_first(
        section2,
        [
            r"Zwroty wskazujące rodzaj zagrożenia\s*(.+?)\s*Zwroty wskazujące środki ostrożności",
            r"Zwrot określający zagrożenie\s*:?\s*(.+?)\s*Zwrot określający środki",
        ],
    )
    # focus on labeling text only
    search_text = compact_text(hazard_block or label_section)

    matches = re.findall(
        r"((?:EUH|H)\d{3}\s*[-–:]?\s*.+?)(?=(?:\b(?:EUH|H)\d{3}\b)|(?:\bP\d{3}\b)|$)",
        search_text,
        flags=re.IGNORECASE,
    )
    hazard_lines = []
    for line in matches:
        cleaned = re.sub(r"\s*·\s*$", "", line).strip()
        if "Zwroty wskazujące środki ostrożności" in cleaned:
            cleaned = cleaned.split("Zwroty wskazujące środki ostrożności", 1)[0].strip()
        hazard_lines.append(cleaned)

    codes = sorted(set(re.findall(r"\bH\d{3}\b", " ".join(hazard_lines))))

    # Only include GHS pictograms explicitly present in section (no H-to-GHS mapping)
    hazard_text = "\n".join(line for line in hazard_lines if line)
    return hazard_text, codes, pictogram_codes


def extract_subsection(section_text: str, labels: list[str], next_labels: list[str]) -> str:
    ascii_text = strip_accents(section_text)
    start_match = None
    for label in labels:
        start_match = re.search(label, ascii_text, flags=re.IGNORECASE)
        if start_match:
            break
    if not start_match:
        return ""

    tail = section_text[start_match.start():]
    tail_ascii = ascii_text[start_match.start():]
    end_positions = []
    for label in next_labels:
        match = re.search(label, tail_ascii, flags=re.IGNORECASE)
        if match and match.start() > 0:
            end_positions.append(match.start())
    snippet = tail[: min(end_positions)] if end_positions else tail
    snippet = re.sub(r"^[^\n:]*:\s*", "", snippet, count=1)
    return normalize_text(snippet)


def extract_first_aid(text: str) -> dict[str, str]:
    ascii_text = strip_accents(text)

    start = None
    for marker in ["4.1. Opis srodkow pierwszej pomocy", "4.1 Opis srodkow pierwszej pomocy", "SEKCJA 4: Srodki pierwszej pomocy"]:
        idx = ascii_text.find(marker)
        if idx >= 0:
            start = idx
            break
    if start is None:
        return {}

    end_candidates = []
    for marker in ["4.2.", "4.2 ", "SEKCJA 5"]:
        idx = ascii_text.find(marker, start + 1)
        if idx >= 0:
            end_candidates.append(idx)
    end = min(end_candidates) if end_candidates else len(text)

    block = text[start:end]
    ascii_block = ascii_text[start:end]

    inline_map = {
        "inhalation": r"Po wdychaniu\s*:?\s*(.+?)(?=Po stycznosci ze skora|Po stycznosci z okiem|Po przelknieciu|4\.2)",
        "skin": r"Po stycznosci ze skora\s*:?\s*(.+?)(?=Po stycznosci z okiem|Po przelknieciu|4\.2)",
        "eyes": r"Po stycznosci z okiem\s*:?\s*(.+?)(?=Po przelknieciu|4\.2)",
        "ingestion": r"Po przelknieciu\s*:?\s*(.+?)(?=4\.2|4\.3|Najwazniejsze ostre)",
    }
    inline_hits: dict[str, str] = {}
    for key, pattern in inline_map.items():
        match = re.search(pattern, ascii_block, flags=re.IGNORECASE | re.DOTALL)
        inline_hits[key] = normalize_text(block[match.start(1):match.end(1)]).strip(" ·") if match else ""
    if any(inline_hits.values()):
        if not inline_hits["ingestion"]:
            fallback_ingestion = capture_first(
                block,
                [r"Po przełknięciu:\s*(.+?)(?=4\.2|4\.3)", r"Po przelknieciu:\s*(.+?)(?=4\.2|4\.3)"],
            )
            inline_hits["ingestion"] = normalize_text(fallback_ingestion).strip(" ·")
        return {
            "inhalation": inline_hits["inhalation"],
            "skin": inline_hits["skin"],
            "eyes": inline_hits["eyes"],
            "ingestion": inline_hits["ingestion"],
        }

    def capture(labels: list[str], next_labels: list[str]) -> str:
        start = None
        matched_label = None
        for label in labels:
            idx = ascii_block.lower().find(label.lower())
            if idx >= 0:
                start = idx
                matched_label = label
                break
        if start is None or matched_label is None:
            return ""

        end = len(block)
        tail_ascii = ascii_block[start + len(matched_label):]
        for label in next_labels:
            idx = tail_ascii.lower().find(label.lower())
            if idx >= 0:
                end = min(end, start + len(matched_label) + idx)

        snippet = block[start:end]
        snippet = snippet.split("\n", 1)
        snippet = snippet[1] if len(snippet) > 1 else ""
        return normalize_text(snippet)

    inhalation = capture(
        ["Przedostanie sie do drog oddechowych", "Drogi oddechowe", "Po wdychaniu"],
        ["Kontakt ze skora", "Kontakt z oczami", "Polkniecie"],
    )
    skin = capture(
        ["Kontakt ze skora", "Po stycznosci ze skora"],
        ["Kontakt z oczami", "Polkniecie"],
    )
    eyes = capture(
        ["Kontakt z oczami", "Po stycznosci z okiem", "Po stycznosci z oczami"],
        ["Polkniecie", "4.2", "4.3"],
    )
    ingestion = capture(
        ["Polkniecie", "Po przelknieciu"],
        ["4.2", "4.3"],
    )
    return {
        "inhalation": inhalation,
        "skin": skin,
        "eyes": eyes,
        "ingestion": ingestion,
    }


def extract_first_aid_general(text: str) -> str:
    section = section_slice(
        text,
        [
            r"4\.2[.\s]+Najważniejsze ostre i opóźnione objawy",
            r"4\.2[.\s]+Najwazniejsze ostre i opoznione objawy",
        ],
        [r"4\.3[.\s]+", r"SEKCJA 5"],
    )
    if not section:
        return ""
    lines = section.splitlines()
    if lines and re.match(r"4\.2[.\s]", lines[0]):
        lines = lines[1:]
    return normalize_text("\n".join(lines))


def extract_fire_data(text: str) -> dict[str, str]:
    section = section_slice(text, [r"SEKCJA 5: Postępowanie w przypadku pożaru"], [r"SEKCJA 6"])
    sub = section_slice(section or text, [r"5\.1[.\s]+Środki gaśnicze"], [r"5\.2[.\s]+", r"5\.3[.\s]+", r"SEKCJA 6"])

    suitable = extract_subsection(
        sub,
        [r"odpowiednie srodki gasnicze", r"środki gaśnicze", r"odpowiednie środki gaśnicze", r"przydatne srodki gasnicze"],
        [r"niewlasciwe srodki gasnicze", r"srodki gasnicze, ktore nie moga byc uzywane", r"środki gaśnicze, które nie mogą być używane", r"nieprzydatne ze wzgledow bezpieczenstwa"],
    )
    unsuitable = extract_subsection(
        sub,
        [r"niewlasciwe srodki gasnicze", r"srodki gasnicze, ktore nie moga byc uzywane", r"środki gaśnicze, które nie mogą być używane", r"nieprzydatne ze wzgledow bezpieczenstwa"],
        [r"5\.2", r"5\.3", r"SEKCJA 6"],
    )

    overview_lines = []
    if suitable:
        suitable = re.sub(r"\s*·\s*Środki gaśnicze\s*$", "", suitable).strip(" ·")
        overview_lines.append(f"Odpowiednie: {compact_text(suitable)}")
    if unsuitable:
        unsuitable = unsuitable.strip(" ·")
        overview_lines.append(f"Nieodpowiednie: {compact_text(unsuitable)}")
    return {
        "overview": "\n".join(overview_lines),
        "suitable": suitable,
        "unsuitable": unsuitable,
    }


def extract_environmental_release(text: str) -> str:
    section6 = section_slice(
        text,
        [r"SEKCJA 6: Postępowanie w przypadku niezamierzonego uwolnienia do środowiska"],
        [r"SEKCJA 7"],
    )
    if not section6:
        return ""
    subsection = section_slice(
        section6,
        [
            r"6\.3[.\s]+Metody i materiałów zapobiegających",
            r"6\.3[.\s]+Metody i materiały zapobiegające",
            r"6\.3[.\s]+",
        ],
        [r"6\.4[.\s]+", r"SEKCJA 7"],
    )
    lines = subsection.splitlines()
    if lines and re.match(r"6\.3[.\s]", lines[0]):
        lines = lines[1:]
    return normalize_text("\n".join(lines))


def extract_protection(text: str) -> dict[str, str]:
    block = section_slice(text, [r"8\.2[.\s]+Kontrola narażenia", r"8\.2[.\s]+Kontrola narazenia"], [r"SEKCJA 9"])
    if not block:
        return {}

    next_labels = [
        r"ochrona rak",
        r"ochrona drog oddechowych",
        r"ochrona skory",
        r"ochrona oczu",
        r"ochrona oczu lub twarzy",
        r"zagrozenia termiczne",
        r"wskazowki dotyczace osobistego osprzetu ochronnego",
    ]

    hands = extract_subsection(block, [r"ochrona rak", r"ochrona skory - ochrona rak"], next_labels)
    respirator = extract_subsection(block, [r"ochrona drog oddechowych"], next_labels)
    skin = extract_subsection(block, [r"ochrona skory", r"ochrona skory - inne"], next_labels)
    eyes = extract_subsection(block, [r"ochrona oczu", r"ochrona oczu lub twarzy"], next_labels)
    if not respirator:
        respirator = capture_first(
            block,
            [r"Ochronę dróg oddechowych\s*:?\s*(.+?)(?=·\s*Zalecane urządzenie filtrujące|·\s*Ochrona rąk|·\s*Ochrona rak|SEKCJA 9)"],
        )
    if not hands:
        hands = capture_first(
            block,
            [r"Ochrona rąk\s*:?\s*(.+?)(?=·\s*Ochronę oczu|·\s*Ochrona ciała|SEKCJA 9)"],
        )
    if not eyes:
        eyes = capture_first(
            block,
            [r"Ochronę oczu lub twarzy\s*:?\s*(.+?)(?=·\s*Ochrona ciała|SEKCJA 9)"],
        )
    if not skin:
        skin = capture_first(
            block,
            [r"Ochrona ciała\s*:?\s*(.+?)(?=SEKCJA 9)", r"Ochrona ciała:([^\n]+)"],
        )
    if hands:
        hands = re.split(r"·\s*Ochronę oczu|·\s*Ochrona ciała", hands, maxsplit=1)[0].strip()
    if respirator:
        respirator = re.split(r"·\s*Zalecane urządzenie filtrujące|·\s*Ochrona rąk", respirator, maxsplit=1)[0].strip()
    if eyes:
        eyes = re.split(r"·\s*Ochrona ciała", eyes, maxsplit=1)[0].strip()
    if skin:
        skin = re.split(r"\*", skin, maxsplit=1)[0].strip()
    return {
        "hands": hands,
        "respirator": respirator,
        "skin": skin,
        "eyes": eyes,
    }
 
def extract_nds(text: str) -> str:
    # Extract NDS (exposure limit) block from sekcja 8.1 up to 8.2
    # slice from section 8.1 to 8.2
    section = section_slice(text, [r"8\.1[.\s]+.*"], [r"8\.2"])
    if not section:
        return ""
    # Look for line containing typical units or 'NDS'
    for line in section.splitlines():
        if "mg/m3" in line or "mg/m³" in line or "NDS" in line:
            # normalize and strip leading label up to first colon
            val = normalize_text(line)
            if ':' in val:
                val = val.split(':', 1)[1].strip()
            return val
    # Fallback: first non-empty line
    for line in section.splitlines():
        if line.strip():
            val = normalize_text(line)
            if ':' in val:
                val = val.split(':', 1)[1].strip()
            return val
    return ""


def extract_work_instructions(text: str) -> dict[str, list[str] | str]:
    section7 = section_slice(
        text,
        [r"SEKCJA 7: Postępowanie z substancjami i mieszaninami oraz ich magazynowanie"],
        [r"SEKCJA 8"],
    )
    handling = section_slice(
        section7 or text,
        [r"7\.1[.\s]+Środki ostrożności dotyczące bezpiecznego postępowania"],
        [r"7\.2[.\s]+", r"7\.3[.\s]+", r"SEKCJA 8"],
    )
    storage = section_slice(
        section7 or text,
        [r"7\.2[.\s]+Warunki bezpiecznego magazynowania"],
        [r"7\.3[.\s]+", r"SEKCJA 8"],
    )
    emergency = section_slice(
        text,
        [r"6\.1[.\s]+Indywidualne środki ostrożności.*?awaryjnych"],
        [r"6\.4[.\s]+", r"SEKCJA 7"],
    )
    section8 = section_slice(
        text,
        [r"8\.2[.\s]+Kontrola narażenia", r"8\.2[.\s]+Kontrola narazenia"],
        [r"SEKCJA 9"],
    )

    def clean_lines(block: str) -> list[str]:
        lines = []
        for raw in block.splitlines():
            line = raw.strip(" -\t")
            if not line:
                continue
            lower = strip_accents(line).lower()
            if re.match(r"^\d+\.\d+", line):
                continue
            if lower in {"patrz:", "zasady higieny:"}:
                continue
            lines.append(line)
        return lines

    handling_lines = clean_lines(handling)
    storage_lines = clean_lines(storage)
    storage_lines = [re.sub(r"^[·\s]*", "", line) for line in storage_lines]
    storage_lines = [line for line in storage_lines if strip_accents(line).lower() not in {"skladowanie:", "niezgodnosci"}]
    emergency_lines = clean_lines(emergency)
    section8_lines = clean_lines(section8)

    before_work: list[str] = []
    during_work: list[str] = []
    after_work: list[str] = []

    for line in handling_lines:
        lower = strip_accents(line).lower()
        if "nie jesc" in lower or "nie pic" in lower or "nie palic" in lower:
            during_work.append(line)
        elif "dobrych praktyk higieny" in lower:
            during_work.append(line)
        elif "przed przerwami" in lower or "po jej zakonczeniu umyc rece" in lower or "umyc rece" in lower:
            after_work.append(line)
        else:
            if line not in before_work:
                before_work.append(line)

    if not during_work:
        for line in handling_lines:
            lower = strip_accents(line).lower()
            if "pracy" in lower or "higieny" in lower:
                during_work.append(line)

    if not after_work:
        for line in handling_lines:
            lower = strip_accents(line).lower()
            if "umyc rece" in lower:
                after_work.append(line)

    if not after_work:
        for line in section8_lines:
            lower = strip_accents(line).lower()
            if "myc rece" in lower or "myc rece przed przerwa" in lower:
                after_work.append("Umyć ręce przed przerwą i po zakończeniu pracy.")
            elif "oczyszczenie skory" in lower or "oczyszczenie skóry" in line.lower():
                after_work.append("Po pracy dokładnie oczyścić skórę.")

    deduped_after: list[str] = []
    for line in after_work:
        if line and line not in deduped_after:
            deduped_after.append(line)
    after_work = deduped_after[:2]

    if emergency_lines:
        before_work.extend([line for line in emergency_lines if line not in before_work][:3])

    if not after_work and emergency_lines:
        for line in emergency_lines:
            lower = strip_accents(line).lower()
            if "wyposazenie ochronne" in lower or "wentylacje" in lower:
                after_work.append(line)
        after_work = after_work[:2]

    if not during_work and before_work:
        during_work = before_work[:2]

    concise_during: list[str] = []
    if any("nie jesc" in strip_accents(line).lower() or "nie pic" in strip_accents(line).lower() or "nie palic" in strip_accents(line).lower() for line in during_work + handling_lines):
        concise_during.append("Nie jeść, nie pić i nie palić w czasie pracy.")
    if any("higien" in strip_accents(line).lower() for line in during_work + handling_lines):
        concise_during.append("Przestrzegać zasad higieny przemysłowej.")
    if any("wentyl" in strip_accents(line).lower() for line in handling_lines):
        concise_during.append("Zapewnić dobrą wentylację stanowiska pracy.")
    if any("zrodla zaplonu" in strip_accents(line).lower() for line in handling_lines):
        concise_during.append("Trzymać z dala od źródeł zapłonu.")
    if any("kontakt z oczami" in strip_accents(line).lower() or "kontaktu z oczami" in strip_accents(line).lower() for line in handling_lines):
        concise_during.append("Unikać kontaktu preparatu z oczami.")
    if any("kontakt z skora" in strip_accents(line).lower() or "kontaktu z skora" in strip_accents(line).lower() or "kontaktu ze skora" in strip_accents(line).lower() for line in handling_lines):
        concise_during.append("Unikać kontaktu preparatu ze skórą.")

    if concise_during:
        during_work = concise_during[:4]

    return {
        "before_work": before_work[:8],
        "during_work": during_work[:6],
        "after_work": after_work[:2],
        "handling": handling_lines[:2],
        "storage": storage_lines[:2],
    }


def extract_data(pdf_path: Path) -> ExtractedData:
    """Wyciągnij dane z karty charakterystyki.

    Ekstrakcja realizowana przez Claude API (`extraction_ai`) — niezależna od
    formatu karty i obsługująca skany (OCR). Stare funkcje `extract_*` /
    `read_pdf_text` pozostają w module jako nieużywany kod referencyjny.
    """
    from extraction_ai import extract_data_ai  # leniwy import — brak cyklu

    return extract_data_ai(pdf_path)


def list_input_pdfs() -> list[Path]:
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)

    pdfs = sorted(INPUT_DIR.glob("*.pdf"))
    if pdfs:
        return pdfs
    return sorted(FALLBACK_INPUT_DIR.glob("*.pdf"))


def resolve_ghs_assets(assets_dir: Path | None = None) -> dict[str, Path]:
    assets_dir = assets_dir or ASSETS_DIR
    ordered_ghs = ["GHS01", "GHS02", "GHS03", "GHS04", "GHS05", "GHS06", "GHS07", "GHS08", "GHS09"]
    mapping: dict[str, Path] = {}

    for code in ordered_ghs:
        direct = assets_dir / f"{code.lower()}.png"
        if direct.exists():
            mapping[code] = direct

    if len(mapping) == len(ordered_ghs):
        return mapping

    extracted = sorted(assets_dir.glob("pikt_*"))
    for code, path in zip(ordered_ghs, extracted):
        mapping.setdefault(code, path)
    return mapping


def resolve_static_assets(assets_dir: Path | None = None, temp_images_dir: Path | None = None) -> dict[str, Path]:
    assets_dir = assets_dir or ASSETS_DIR
    temp_images_dir = temp_images_dir or TEMP_IMAGES_DIR
    candidates = {
        "gloves": ["gloves.png", "rekawice.png", "template_image2.png", "main_r16c1_h967002ea.png"],
        "respirator": ["respirator.png", "maska.png", "template_image3.png", "main_r16c5_h72827d92.png"],
        "clothing": ["clothing.png", "kombinezon.png", "template_image4.png", "main_r17c1_h8a7a25f5.png"],
        "goggles": ["goggles.png", "okulary.png", "template_image7.png", "main_r17c5_h14d354db.png"],
        "no_smoking": ["no_smoking.png", "zakaz_palenia.png", "template_image6.jpeg", "main_r29c1_h6bdb16fd.png"],
        "no_food": ["no_food.png", "zakaz_jedzenia.png", "template_image5.jpeg", "main_r31c1_h4d231887.png"],
    }
    resolved: dict[str, Path] = {}
    for key, names in candidates.items():
        for name in names:
            for base in (assets_dir, temp_images_dir):
                path = base / name
                if path.exists():
                    resolved[key] = path
                    break
            if key in resolved:
                break
    return resolved


def add_image(
    ws,
    image_path: Path,
    anchor: str | OneCellAnchor,
    width: int | None = None,
    height: int | None = None,
) -> None:
    if not image_path.exists():
        return
    try:
        img = Image(str(image_path))
    except (FileNotFoundError, OSError, UnidentifiedImageError):
        return
    if width:
        img.width = width
    if height:
        img.height = height
    img.anchor = anchor
    ws.add_image(img)


def _strip_placeholder(value: str | None) -> str | None:
    if not value:
        return value
    new, hits = re.subn(
        r"\s*z\s+pkt\.?\s*\d+(?:\.\d+)?\s*KCH", "", value, flags=re.IGNORECASE
    )
    if hits:
        # Usuń „wiszący” ogon po placeholderze, np. „Drogi oddechowe: . ” -> „Drogi oddechowe:”
        new = re.sub(r"[\s.]+$", "", new)
    return new.strip()


def _combine_value(existing, new: str) -> str:
    parts = [p for p in (existing, new) if p]
    return "\n".join(parts)


def _get_merged_parent_value(ws, address: str) -> str | None:
    for merged_range in ws.merged_cells.ranges:
        if address in merged_range:
            parent = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return parent.value
    return None


def _is_bold(cell) -> bool:
    return bool(cell.font and cell.font.bold)


def _rich_append(existing: str | None, new: str) -> CellRichText:
    parts: list[TextBlock] = []
    if existing:
        header_font = InlineFont(b=True)
        header_font.rFont = "Arial"
        parts.append(TextBlock(header_font, existing))
    if new:
        prefix = "\n" if existing else ""
        body_font = InlineFont(b=False)
        body_font.rFont = "Arial"
        parts.append(TextBlock(body_font, prefix + new))
    return CellRichText(*parts)


def write_value(ws, address: str, value: str, append: bool = False) -> None:
    value = _clean_xml(value)
    target = ws[address]
    if target.__class__.__name__ == "MergedCell":
        for merged_range in ws.merged_cells.ranges:
            if address in merged_range:
                cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                if append:
                    parent_value = _strip_placeholder(cell.value)
                    if _is_bold(cell) and (parent_value or value):
                        cell.value = _rich_append(parent_value, value)
                    else:
                        cell.value = _combine_value(parent_value, value)
                else:
                    cell.value = value
                ensure_visible_style(cell)
                return
    if append:
        existing = _strip_placeholder(target.value)
        if _is_bold(target) and (existing or value):
            target.value = _rich_append(existing, value)
        else:
            target.value = _combine_value(existing, value)
    else:
        target.value = value
    ensure_visible_style(target)


def ensure_visible_style(cell) -> None:
    font = copy(cell.font)
    if font.color is None or getattr(font.color, "type", None) == "rgb":
        font.color = "FF000000"
    cell.font = font
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    if cell.coordinate in {"B7", "D7", "F7"}:
        alignment.horizontal = "center"
        alignment.vertical = "center"
    else:
        alignment.horizontal = alignment.horizontal or "left"
        alignment.vertical = "top"
    cell.alignment = alignment
    if cell.coordinate in {"C29", "C30", "C31", "C32", "C33", "C34"}:
        font = copy(cell.font)
        font.size = 9
        cell.font = font


_AUTOFIT_LINE_PT = 12.6        # ~jedna linia tekstu Arial 10 pt
_AUTOFIT_PAD_PT = 2.0          # zapas pionowy na komórkę
_AUTOFIT_CHAR_FUDGE = 0.9      # ułamek „znaków szerokości" realnie mieszczących się w linii
_DEFAULT_COL_CHARS = 8.43
GHS_BLOCK_ROW_PT = 40.0        # min. wysokość wiersza mieszczącego piktogram GHS 1,3 cm (~49 px)


def autofit_row_heights(ws, addresses) -> None:
    """Zwiększ wysokość wierszy, by zmieścił się zawinięty tekst wpisanych komórek.

    Dotyczy wyłącznie komórek faktycznie wypełnianych danymi (`addresses`) oraz
    ich scalonych bloków — statyczna treść szablonu (nagłówki, instrukcje pracy)
    pozostaje nietknięta. Wysokości są tylko zwiększane (nigdy poniżej szablonu):
    krótka treść zachowuje geometrię `wzor.xlsx`, a długa treść z karty sama się
    rozsuwa — bez ręcznego poprawiania komórek po wygenerowaniu.
    """
    from openpyxl.utils import coordinate_to_tuple, get_column_letter

    def col_chars(col: int) -> float:
        width = ws.column_dimensions[get_column_letter(col)].width
        return width if width else _DEFAULT_COL_CHARS

    required: dict[int, float] = {}
    seen: set[tuple[int, int]] = set()
    for address in addresses:
        row, col = coordinate_to_tuple(address)
        r0, r1, c0, c1 = row, row, col, col
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                r0, r1, c0, c1 = mr.min_row, mr.max_row, mr.min_col, mr.max_col
                break
        if (r0, c0) in seen:
            continue
        seen.add((r0, c0))
        value = ws.cell(row=r0, column=c0).value
        if value is None:
            continue
        usable = max(sum(col_chars(c) for c in range(c0, c1 + 1)) * _AUTOFIT_CHAR_FUDGE, 1.0)
        lines = 0
        for segment in str(value).split("\n"):
            seg = segment.strip()
            lines += max(1, math.ceil(len(seg) / usable)) if seg else 1
        per_row = (lines * _AUTOFIT_LINE_PT + _AUTOFIT_PAD_PT) / (r1 - r0 + 1)
        for r in range(r0, r1 + 1):
            if per_row > required.get(r, 0.0):
                required[r] = per_row

    for r, need in required.items():
        current = ws.row_dimensions[r].height or 0.0
        if need > current:
            ws.row_dimensions[r].height = round(need, 1)


def _fit_ghs_block(ws, n_pictograms: int) -> None:
    """Zapewnij wysokość wierszy 11..14 na piktogramy GHS (2 na wiersz, 40 px)."""
    if n_pictograms <= 0:
        return
    rows_used = min(math.ceil(n_pictograms / 2), 4)
    for r in range(11, 11 + rows_used):
        if (ws.row_dimensions[r].height or 0.0) < GHS_BLOCK_ROW_PT:
            ws.row_dimensions[r].height = GHS_BLOCK_ROW_PT


EMU_PER_PIXEL = 9525
EMU_PER_CM = 360000
STATIC_IMAGE_MIN_MARGIN_PX = 3.0


def _cm_to_px(cm: float) -> int:
    return int(round(cm / 2.54 * 96))


# Rozmiary piktogramów wg uwag recenzenta (kotwica wyśrodkowana w komórce):
# - środki ochrony (rękawice, maska, odzież, okulary): 2,0 × 2,0 cm
# - zakazy (zakaz palenia / jedzenia): 1,75 × 1,75 cm
# - piktogramy GHS (zagrożenia, blok A11): 1,3 × 1,3 cm
PROTECTION_IMAGE_SIZE_CM = 2.0
PROHIBITION_IMAGE_SIZE_CM = 1.75
STATIC_IMAGE_SIZES_CM = {
    "gloves": PROTECTION_IMAGE_SIZE_CM,
    "respirator": PROTECTION_IMAGE_SIZE_CM,
    "clothing": PROTECTION_IMAGE_SIZE_CM,
    "goggles": PROTECTION_IMAGE_SIZE_CM,
    "no_smoking": PROHIBITION_IMAGE_SIZE_CM,
    "no_food": PROHIBITION_IMAGE_SIZE_CM,
}

GHS_IMAGE_SIZE_CM = 1.3
GHS_IMAGE_SIZE_PX = _cm_to_px(GHS_IMAGE_SIZE_CM)   # ~49 px


def _col_width_px(ws, col_letter: str) -> float:
    width = ws.column_dimensions[col_letter].width or _DEFAULT_COL_CHARS
    return width * 7.0 + 5.0


def _row_height_px(ws, row: int) -> float:
    height = ws.row_dimensions[row].height or (ws.sheet_format.defaultRowHeight or 15.0)
    return height * 96.0 / 72.0


def static_image_anchor(ws, address: str, size_cm: float) -> OneCellAnchor:
    """Kotwica komórkowa wyśrodkowująca piktogram `size_cm` cm w jego komórce.

    Offset pionowy/poziomy dosuwa ikonę do środka komórki (z minimalnym
    marginesem), dzięki czemu nie nachodzi na sąsiednie wiersze ani nagłówki.
    Gdy komórka jest niższa niż ikona (np. bloki zakazów), zostaje sam margines
    górny — ikona mieści się w scalonym bloku, nie dotykając wiersza powyżej.
    """
    from openpyxl.utils import coordinate_to_tuple, get_column_letter

    size_px = _cm_to_px(size_cm)
    size_emu = int(round(size_cm * EMU_PER_CM))
    row, col = coordinate_to_tuple(address)
    col_letter = get_column_letter(col)
    off_x = max((_col_width_px(ws, col_letter) - size_px) / 2.0, STATIC_IMAGE_MIN_MARGIN_PX)
    off_y = max((_row_height_px(ws, row) - size_px) / 2.0, STATIC_IMAGE_MIN_MARGIN_PX)
    marker = AnchorMarker(
        col=col - 1,
        row=row - 1,
        colOff=int(off_x * EMU_PER_PIXEL),
        rowOff=int(off_y * EMU_PER_PIXEL),
    )
    ext = XDRPositiveSize2D(size_emu, size_emu)
    return OneCellAnchor(_from=marker, ext=ext)


def layout_ghs_images(ws, pictograms: list[str]) -> list[tuple[str, OneCellAnchor | str, int]]:
    """Rozmieść piktogramy GHS w bloku zagrożeń A11:G14.

    Stały rozmiar, prosty anchor komórkowy, kolumny F/G × wiersze 11..14
    (max 8). NIE przelicza pikseli z szerokości kolumn ani nie zmienia
    wysokości wierszy — układ pozostaje wierny szablonowi `wzor.xlsx`.
    """
    cols = ["F", "G"]
    rows = [11, 12, 13, 14]
    result: list[tuple[str, OneCellAnchor | str, int]] = []
    for i, code in enumerate(pictograms[: len(cols) * len(rows)]):
        anchor = f"{cols[i % len(cols)]}{rows[i // len(cols)]}"
        result.append((code, anchor, GHS_IMAGE_SIZE_PX))
    return result


def populate_workbook(
    data: ExtractedData,
    output_path: Path,
    template_path: Path | None = None,
    assets_dir: Path | None = None,
    temp_images_dir: Path | None = None,
) -> None:
    template_path = template_path or TEMPLATE_PATH
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # Czcionka 11 w sekcjach czynności (tekst statyczny szablonu) — wg uwag.
    for addr in SECTION_FONT11_CELLS:
        cell = ws[addr]
        if cell.__class__.__name__ == "MergedCell":
            continue
        section_font = copy(cell.font)
        section_font.size = 11
        cell.font = section_font

    written: list[str] = []

    def put(key: str, value: str, append: bool = False) -> None:
        addr = CELL_MAP[key]
        write_value(ws, addr, value, append=append)
        written.append(addr)

    put("producer", data.producer)
    put("product_name", data.product_name)
    put("revision_date", data.revision_date)
    # A11 (scalona A11:G14): wpisz zwroty H tylko gdy są — w przeciwnym razie
    # zostaw placeholder/format szablonu nienaruszony.
    if (data.hazard_text or "").strip():
        put("hazards", data.hazard_text)
    put("hand_protection", data.hand_protection, append=True)
    put("respiratory_protection", data.respiratory_protection, append=True)
    put("skin_protection", data.skin_protection, append=True)
    put("eye_protection", data.eye_protection, append=True)
    # A40 „Uwagi ogólne" — stały tekst, BEZ danych z karty (wg uwag recenzenta).
    put("first_aid_general", FIRST_AID_GENERAL_TEXT, append=True)
    put("first_aid_inhalation", data.first_aid_inhalation, append=True)
    put("first_aid_skin", data.first_aid_skin, append=True)
    put("first_aid_eyes", data.first_aid_eyes, append=True)
    put("first_aid_ingestion", data.first_aid_ingestion, append=True)
    put("fire_overview", data.fire_overview, append=True)
    put("fire_suitable", data.fire_suitable, append=True)
    put("fire_unsuitable", data.fire_unsuitable, append=True)
    put("environmental_release", data.environmental_release, append=True)
    # A54/A55 — stałe formułki (postępowanie z produktem, magazynowanie).
    put("handling_1", HANDLING_TEXT, append=True)
    put("storage_1", STORAGE_TEXT, append=True)

    ws._images = []
    static_assets = resolve_static_assets(assets_dir=assets_dir, temp_images_dir=temp_images_dir)
    for key, anchor in STATIC_IMAGE_ANCHORS.items():
        size_cm = STATIC_IMAGE_SIZES_CM.get(key, PROTECTION_IMAGE_SIZE_CM)
        size_px = _cm_to_px(size_cm)
        add_image(
            ws,
            static_assets.get(key, Path()),
            static_image_anchor(ws, anchor, size_cm),
            width=size_px,
            height=size_px,
        )

    ghs_assets = resolve_ghs_assets(assets_dir=assets_dir)
    placed_ghs = 0
    for code, anchor, size in layout_ghs_images(ws, data.hazard_pictograms):
        asset = ghs_assets.get(code)
        if asset:
            add_image(ws, asset, anchor, width=size, height=size)
            placed_ghs += 1

    # Dopasuj wysokość wierszy do treści: długi tekst rośnie sam, krótki
    # zachowuje geometrię szablonu (wysokości tylko rosną, nigdy nie maleją).
    autofit_row_heights(ws, written)
    _fit_ghs_block(ws, placed_ghs)

    # add separate sheet for NDS (exposure limit) from section 8.1
    nds_ws = wb.create_sheet(title="NDS")
    # use extracted NDS text directly
    nds_value = _clean_xml(data.nds or "")
    nds_ws["A1"] = nds_value
    ensure_visible_style(nds_ws["A1"])

    wb.save(output_path)
    wb.close()


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Brak szablonu: {TEMPLATE_PATH}")
    if not ASSETS_DIR.exists():
        raise FileNotFoundError(f"Brak folderu assets: {ASSETS_DIR}")

    pdfs = list_input_pdfs()
    if not pdfs:
        print("Brak plików PDF w DANE_WEJSCIOWE ani w DOCS.")
        return

    for pdf_path in pdfs:
        data = extract_data(pdf_path)
        product = data.product_name or pdf_path.stem
        output_name = f"Instrukcja_BHP_{sanitize_filename(product)}.xlsx"
        output_path = OUTPUT_DIR / output_name
        populate_workbook(data, output_path)
        print(f"Zapisano: {output_path.name}")


if __name__ == "__main__":
    main()
